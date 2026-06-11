"""Batch job status, results, export, and progress-stream endpoints."""

from __future__ import annotations

import asyncio
import csv
import io
import json
from collections.abc import AsyncIterator
from typing import Literal

from fastapi import APIRouter, HTTPException, Query, Response, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import (
    LABEL_FIELD_NAMES,
    BatchProgress,
    BatchSummary,
    JobResultsResponse,
    JobState,
    JobStatusResponse,
    OverallStatus,
    VerificationResult,
)
from batch import store
from batch.orchestrator import start_batch

router = APIRouter(prefix="/jobs", tags=["jobs"])

#: How often a reconnecting SSE client polls for newly-completed labels.
_POLL_INTERVAL_SECONDS = 0.1

#: The six TTB-required comparison fields, excluding "government_warning"
#: (which has its own valid/invalid status rather than a MatchStatus).
_COMPARISON_FIELDS = LABEL_FIELD_NAMES[:-1]

#: Per-label export columns (ISSUE 3.5): filename and overall outcome, then
#: extracted/expected/status/confidence for each comparison field, then the
#: Government Warning's extracted/expected text and valid/invalid status.
EXPORT_HEADER: list[str] = [
    "filename",
    "overall_status",
    "confidence_score",
    "image_quality_score",
    *(
        column
        for name in _COMPARISON_FIELDS
        for column in (f"{name}_extracted", f"{name}_expected", f"{name}_status", f"{name}_confidence")
    ),
    "government_warning_extracted",
    "government_warning_expected",
    "government_warning_status",
]

#: Columns whose values are status strings, eligible for color-coding in the
#: .xlsx export (MatchStatus / OverallStatus / "valid" / "invalid").
_STATUS_COLUMNS = frozenset(
    ["overall_status", *(f"{name}_status" for name in _COMPARISON_FIELDS), "government_warning_status"]
)

#: Excel's standard "Good"/"Neutral"/"Bad" conditional-formatting palette
#: (fill, font), keyed by the status value it applies to.
_STATUS_FILLS: dict[str, tuple[str, str]] = {
    "MATCH": ("FFC6EFCE", "FF006100"),
    "PARTIAL_MATCH": ("FFFFEB9C", "FF9C6500"),
    "PARTIAL": ("FFFFEB9C", "FF9C6500"),
    "NO_MATCH": ("FFFFC7CE", "FF9C0006"),
    "FAIL": ("FFFFC7CE", "FF9C0006"),
    "ERROR": ("FFFFC7CE", "FF9C0006"),
    "valid": ("FFC6EFCE", "FF006100"),
    "invalid": ("FFFFC7CE", "FF9C0006"),
}

#: Header row styling for the .xlsx export — white text on the Treasury
#: brand color (treasury-700, see frontend/src/index.css).
_HEADER_FILL = PatternFill("solid", fgColor="FF0B5D44")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")


def _require_job(job_id: str) -> store.Job:
    job = store.get_job(job_id)
    if job is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"job '{job_id}' not found",
        )
    return job


def _completed_results(job: store.Job) -> list[VerificationResult]:
    """Results filled in so far, skipping `None` placeholders for in-flight labels."""
    return [r for r in job.results if r is not None]


def _summarize(job: store.Job) -> BatchSummary:
    summary = BatchSummary()
    for result in _completed_results(job):
        if result.overall_status is OverallStatus.MATCH:
            summary.match += 1
        elif result.overall_status is OverallStatus.PARTIAL:
            summary.partial += 1
        elif result.overall_status is OverallStatus.FAIL:
            summary.fail += 1
        else:
            summary.error += 1
    return summary


@router.get("/{job_id}/status", response_model=JobStatusResponse)
def job_status(job_id: str) -> JobStatusResponse:
    job = _require_job(job_id)
    return JobStatusResponse(
        job_id=job.job_id,
        state=job.state,
        completed=job.completed,
        total=job.total,
    )


@router.get("/{job_id}/results", response_model=JobResultsResponse)
def job_results(job_id: str) -> JobResultsResponse:
    job = _require_job(job_id)
    return JobResultsResponse(
        job_id=job.job_id,
        state=job.state,
        summary=_summarize(job),
        results=_completed_results(job),
    )


def _export_row(result: VerificationResult) -> list[str]:
    """Build one export row, in `EXPORT_HEADER` order, from a single result."""
    by_field = {fc.field: fc for fc in result.fields}
    row: list[str] = [
        result.filename or "",
        result.overall_status.value,
        f"{result.confidence_score:.1f}",
        f"{result.image_quality_score:.1f}",
    ]
    for name in _COMPARISON_FIELDS:
        fc = by_field.get(name)
        if fc is None:
            row += ["", "", "", ""]
        else:
            row += [fc.extracted or "", fc.expected or "", fc.status.value, f"{fc.score:.1f}"]
    warning = result.government_warning
    row += [
        warning.extracted_text or "",
        warning.expected_text or "",
        "valid" if warning.valid else "invalid",
    ]
    return row


def _export_csv(job: store.Job) -> Response:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(EXPORT_HEADER)
    for result in _completed_results(job):
        writer.writerow(_export_row(result))
    return Response(
        content=buffer.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="results_{job.job_id}.csv"'},
    )


def _column_width(name: str) -> int:
    if name == "overall_status" or name.endswith(("_status", "_confidence", "_score")):
        return 14
    return 28


def _export_xlsx(job: store.Job) -> Response:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    for col, name in enumerate(EXPORT_HEADER, start=1):
        cell = sheet.cell(row=1, column=col, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        sheet.column_dimensions[get_column_letter(col)].width = _column_width(name)
    sheet.freeze_panes = "A2"

    for row_index, result in enumerate(_completed_results(job), start=2):
        for col, (name, value) in enumerate(zip(EXPORT_HEADER, _export_row(result), strict=True), start=1):
            cell = sheet.cell(row=row_index, column=col, value=value)
            if name in _STATUS_COLUMNS and value in _STATUS_FILLS:
                fill_color, font_color = _STATUS_FILLS[value]
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.font = Font(color=font_color)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="results_{job.job_id}.xlsx"'},
    )


@router.get(
    "/{job_id}/export",
    summary="Export batch results as CSV or Excel",
    responses={
        200: {
            "content": {
                "text/csv": {},
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {},
            },
            "description": "RFC 4180 CSV (default, `format=csv`) or a formatted .xlsx workbook (`format=xlsx`)",
        }
    },
)
def job_export(job_id: str, format: Literal["csv", "xlsx"] = Query("csv")) -> Response:
    job = _require_job(job_id)
    return _export_xlsx(job) if format == "xlsx" else _export_csv(job)


def _sse(event: str, data: dict) -> str:
    return f"event: {event}\ndata: {json.dumps(data)}\n\n"


def _progress_event(progress: BatchProgress) -> str:
    event = "error" if progress.latest.overall_status is OverallStatus.ERROR else "progress"
    return _sse(event, progress.model_dump(mode="json"))


def _complete_event(job: store.Job) -> str:
    return _sse(
        "complete",
        {
            "job_id": job.job_id,
            "state": job.state.value,
            "completed": job.completed,
            "total": job.total,
            "summary": _summarize(job).model_dump(),
        },
    )


async def _stream_progress(job: store.Job) -> AsyncIterator[str]:
    """Emit `progress`/`error` events as labels finish, then `complete`.

    The first connection for a PENDING job drives the orchestrator directly,
    so the whole batch runs to completion within this one response. A
    reconnecting `EventSource` (job already PROCESSING or COMPLETED) instead
    polls `job.results` for labels it hasn't seen yet, replaying any it
    missed before catching up to new ones.
    """
    if job.state is JobState.PENDING:
        labels, job.labels = job.labels, []
        async for progress in start_batch(job.job_id, labels):
            yield _progress_event(progress)
    else:
        seen: set[int] = set()
        while True:
            for index, result in enumerate(job.results):
                if result is None or index in seen:
                    continue
                seen.add(index)
                yield _progress_event(
                    BatchProgress(job_id=job.job_id, completed=len(seen), total=job.total, latest=result)
                )
            if job.state is JobState.COMPLETED:
                break
            await asyncio.sleep(_POLL_INTERVAL_SECONDS)

    yield _complete_event(job)


@router.get(
    "/{job_id}/stream",
    summary="Stream batch progress via Server-Sent Events",
    responses={200: {"content": {"text/event-stream": {}}, "description": "progress/error/complete SSE events"}},
)
async def job_stream(job_id: str) -> StreamingResponse:
    job = _require_job(job_id)
    return StreamingResponse(
        _stream_progress(job),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )
