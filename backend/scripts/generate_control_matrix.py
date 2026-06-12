"""Generate `docs/fedramp/CONTROL-MATRIX.xlsx` (ISSUE 4.5, AC5).

Renders the 33-control NIST SP 800-53 Rev. 5 Moderate-baseline matrix —
the FedRAMP Control Coverage Matrix from `project-management/PROJECT-PLAN.md`
plus the four PoC-specific additions (AU-14, SI-11, CP-10, PL-8) — as a
color-coded spreadsheet. Status values and implementation notes mirror
`docs/fedramp/SSP-final.md` §8 exactly.

Run from the `backend/` directory:

    python scripts/generate_control_matrix.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(__file__).resolve().parents[2] / "docs" / "fedramp" / "CONTROL-MATRIX.xlsx"

HEADERS = ["NIST Control ID", "Control Name", "Family", "Status", "Phase / Issue", "Implementation Notes / Evidence"]

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)

STATUS_FILLS = {
    "Implemented": PatternFill("solid", fgColor="C6EFCE"),
    "Planned": PatternFill("solid", fgColor="FFEB9C"),
    "Not Applicable": PatternFill("solid", fgColor="D9D9D9"),
    "Inherited": PatternFill("solid", fgColor="DDEBF7"),
}

WRAP_TOP_LEFT = Alignment(wrap_text=True, vertical="top", horizontal="left")

# (control_id, name, family, status, phase/issue, implementation notes/evidence)
CONTROLS = [
    ("AC-3", "Access Enforcement", "Access Control", "Implemented", "ISSUE 3.7",
     "Every /jobs/* request requires a valid signed session cookie (403 otherwise); each batch "
     "job is scoped to the session that created it (404 for any other session). See "
     "SESSION-MANAGEMENT.md."),
    ("AC-4", "Information Flow Enforcement", "Access Control", "Implemented", "-",
     "All processing is in-memory within the authorization boundary; the only cross-boundary "
     "flow is the single whitelisted Claude Vision endpoint, gated by OCR_MODE."),
    ("AC-17", "Remote Access", "Access Control", "Implemented", "-",
     "The frontend is reachable only over HTTPS; /api/* is same-origin via the nginx reverse "
     "proxy (no CORS). TLS termination at the network ingress is inherited from the hosting GSS."),
    ("AU-2", "Event Logging", "Audit and Accountability", "Implemented", "ISSUE 2.7",
     "backend/app/audit.py configures structlog to emit one JSON object per line. Events: "
     "request_received, request_completed, ocr_started, ocr_completed, match_completed, "
     "request_error, session_expired."),
    ("AU-3", "Content of Audit Records", "Audit and Accountability", "Implemented", "-",
     "request_completed records timestamp, request_id, endpoint, method, status_code, "
     "duration_ms, session_id, and ocr_engine_used. request_error records status_code, error, "
     "and message."),
    ("AU-9", "Protection of Audit Information", "Audit and Accountability", "Implemented", "-",
     "Logs go to stdout only (structlog.PrintLoggerFactory) -- never to a file inside the "
     "container. Helper-function signatures structurally exclude PII."),
    ("AU-14", "Session Audit", "Audit and Accountability", "Implemented", "ISSUE 4.2",
     "The 300-label load test drives one batch end-to-end through the same session-cookie flow "
     "used by the browser, confirming request_completed/ocr_completed audit events are emitted "
     "correctly for every label under realistic batch load."),
    ("CM-2", "Baseline Configuration", "Configuration Management", "Implemented", "-",
     "All runtime dependencies are pinned (backend/requirements.txt, frontend lockfile); base "
     "images are digest-pinned (docker/*.Dockerfile)."),
    ("CM-3", "Configuration Change Control", "Configuration Management", "Implemented", "-",
     "GitHub branch protection on main requires the aggregate CI Success status check; all "
     "changes land via reviewed PRs (CODEOWNERS)."),
    ("CM-6", "Configuration Settings", "Configuration Management", "Implemented", "-",
     "All tunables are environment variables documented in .env.example (LOG_LEVEL, OCR_MODE, "
     "MAX_IMAGE_MB, MAX_BATCH_MB, SESSION_TTL_HOURS, SESSION_SECRET_KEY, etc.) -- no hardcoded "
     "configuration or secrets."),
    ("CM-7", "Least Functionality", "Configuration Management", "Implemented", "-",
     "Backend container runs as an unprivileged app user; both images are based on "
     "-slim/-alpine variants with minimal installed packages."),
    ("IA-2", "Identification and Authentication", "Identification and Authentication", "Implemented", "ISSUE 3.7",
     "Each browser is identified by a cryptographically random session id "
     "(secrets.token_urlsafe(32)), issued on first visit and validated on every request via an "
     "HMAC-SHA256-signed cookie."),
    ("SC-8", "Transmission Confidentiality and Integrity", "System and Communications Protection", "Implemented", "-",
     "The anthropic SDK uses HTTPS (TLS 1.2+) for the Claude Vision call; reviewer-facing HTTPS "
     "and ingress TLS termination are inherited from the hosting GSS. The internal "
     "frontend<->backend hop (TB-1) is plain HTTP but never leaves the authorization boundary."),
    ("SC-23", "Session Authenticity", "System and Communications Protection", "Implemented", "ISSUE 3.7",
     "The session-id cookie is HttpOnly, Secure, SameSite=Strict, HMAC-SHA256-signed, and "
     "expires (Max-Age) in lockstep with the server-side session TTL."),
    ("SC-28", "Protection of Information at Rest", "System and Communications Protection", "Implemented", "-",
     "No disk writes of label images, application data, extracted fields, or match results "
     "anywhere in backend/app, backend/ocr, or backend/matching -- all state lives in "
     "backend/batch/store.py's in-memory dict."),
    ("SI-3", "Malicious Code Protection", "System and Information Integrity", "Implemented", "ISSUE 2.6",
     "Bandit (Python SAST) and Trivy (container image scanning, CRITICAL gate) run on every PR; "
     "results in SAST-RESULTS.md."),
    ("SI-7", "Software, Firmware, and Information Integrity", "System and Information Integrity", "Implemented", "ISSUE 2.5",
     "backend/matching/exact_validator.py performs a word-for-word, ALL-CAPS-prefix exact match "
     "of the Government Warning text against the application data."),
    ("SI-10", "Information Input Validation", "System and Information Integrity", "Implemented", "ISSUE 3.6, 4.1",
     "All API I/O is typed via Pydantic models; both /verify and /verify/batch validate every "
     "image by magic-byte signature and size (HTTP 413/415); application_csv rejects unknown "
     "columns (HTTP 422); a 20+ case fuzz suite confirms malformed input never produces a 5xx. "
     "Preprocessing (deskew/denoise/sharpen/contrast) degrades safely on failure."),
    ("SI-11", "Error Handling", "System and Information Integrity", "Implemented", "-",
     "A uniform ErrorResponse{error, message, request_id} envelope is returned for all HTTP and "
     "validation errors, and every error response also emits a request_error audit event."),
    ("SI-12", "Information Management and Retention", "System and Information Integrity", "Implemented", "ISSUE 3.5",
     "SESSION_TTL_HOURS (default 4h) bounds every batch job: "
     "backend/batch/store.py::_reap_expired drops idle jobs and emits session_expired. Verified "
     "by backend/tests/test_session_store.py."),
    ("SI-16", "Memory Protection", "System and Information Integrity", "Implemented", "ISSUE 3.6",
     "Per-image (MAX_IMAGE_MB, default 20MB) and per-batch (MAX_BATCH_MB, default 500MB) size "
     "limits are enforced before any image is processed (HTTP 413)."),
    ("SI-17", "Fail-Safe Procedures", "System and Information Integrity", "Implemented", "ISSUE 4.4",
     "A catch-all Exception handler guarantees every unhandled error returns the ErrorResponse "
     "envelope. backend/app/pipeline.py::run_verification returns plain-language ERROR results "
     "instead of crashing; the batch orchestrator isolates per-label failures; the frontend "
     "ErrorBoundary and useJobStream reconnecting state cover client-side failures."),
    ("CP-10", "System Recovery and Reconstitution", "Contingency Planning", "Implemented", "ISSUE 4.2",
     "The 300-label load test demonstrated that when Claude Vision becomes unavailable "
     "mid-batch, backend/ocr/adapter.py::extract_fields automatically fails over to local "
     "Tesseract OCR -- the batch completes with zero crashes and no operator intervention."),
    ("PL-2", "System Security Plan", "Planning", "Implemented", "ISSUE 1.6 -> 4.5",
     "This SSP (SSP-final.md), drafted in Phase 1 (SSP-draft.md) and finalized in Phase 4 with "
     "the complete control matrix, POAM.md, DATA-FLOW-final.md, and INCIDENT-RESPONSE-PLAN.md."),
    ("PL-8", "Security and Privacy Architectures", "Planning", "Implemented", "ISSUE 4.3",
     "The frontend meets WCAG 2.1 AA: an automated axe-core scan finds zero violations across "
     "all 3 routes, and a manual keyboard-navigation + accessibility-tree pass confirmed a "
     "logical tab order, visible focus indicators, correct landmark/heading/label structure, "
     "and aria-live announcements. See ACCESSIBILITY-REPORT.md."),
    ("RA-2", "Security Categorization", "Risk Assessment", "Implemented", "-",
     "FIPS 199 categorization performed in SSP-final.md §2 -- Confidentiality, Integrity, and "
     "Availability are all Moderate, driving selection of the FedRAMP Moderate control "
     "baseline."),
    ("RA-3", "Risk Assessment", "Risk Assessment", "Planned", "ISSUE 4.6 (#48)",
     "A dedicated threat model and attack-surface analysis (THREAT-MODEL.md) is scoped for "
     "ISSUE 4.6 -- Threat Model Documentation. This is the sole open item tracked in POAM.md."),
    ("RA-5", "Vulnerability Monitoring and Scanning", "Risk Assessment", "Implemented", "ISSUE 2.6, 4.5",
     "pip-audit (Python SCA), npm audit (Node SCA), and Trivy (container image CVE scanning) "
     "run on every PR; results in SAST-RESULTS.md, re-run for this final package on 2026-06-11."),
    ("SA-9", "External System Services", "System and Services Acquisition", "Implemented", "-",
     "The Claude Vision API is the only external service ALVA depends on, and it is optional: "
     "explicitly whitelisted, gated by OCR_MODE/ANTHROPIC_API_KEY, with automatic fail-over to "
     "local Tesseract."),
    ("SA-11", "Developer Testing and Evaluation", "System and Services Acquisition", "Implemented", "-",
     "Every PR runs the backend pytest suite, the frontend vitest + jest-axe accessibility "
     "suite, bandit SAST, eslint-plugin-security, pip-audit, and npm audit, gated by the "
     "required CI Success status check on main."),
    ("CA-5", "Plan of Action and Milestones", "Security Assessment and Authorization", "Implemented", "ISSUE 4.5",
     "Known gaps are tracked with remediation timelines in POAM.md; as of this matrix the sole "
     "open item is RA-3 (Threat Model, ISSUE 4.6)."),
    ("CA-7", "Continuous Monitoring", "Security Assessment and Authorization", "Implemented", "-",
     "The CI pipeline re-runs the full SAST/SCA/container-scan suite on every PR to main, gated "
     "by the aggregate CI Success required status check -- control effectiveness is "
     "continuously re-verified, not assessed once."),
    ("IR-8", "Incident Response Plan", "Incident Response", "Implemented", "ISSUE 4.5",
     "INCIDENT-RESPONSE-PLAN.md defines detection, triage, and response procedures for "
     "OCR/processing failures, data exposure events, and availability incidents."),
]

COLUMN_WIDTHS = [16, 38, 36, 16, 18, 90]


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Control Matrix"

    ws.append(HEADERS)
    for col, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_TOP_LEFT
    ws.freeze_panes = "A2"

    for row in CONTROLS:
        ws.append(row)
        row_idx = ws.max_row
        status = row[3]
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = WRAP_TOP_LEFT
            cell.fill = STATUS_FILLS[status]

    implemented = sum(1 for row in CONTROLS if row[3] == "Implemented")
    planned = sum(1 for row in CONTROLS if row[3] == "Planned")

    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["NIST SP 800-53r5 Moderate Baseline -- ALVA Control Matrix (ISSUE 4.5)"])
    summary_ws.append([])
    summary_ws.append(["Total controls", len(CONTROLS)])
    summary_ws.append(["Implemented", implemented])
    summary_ws.append(["Planned", planned])
    summary_ws.append(["Not Applicable", sum(1 for row in CONTROLS if row[3] == "Not Applicable")])
    summary_ws.append(["Inherited", sum(1 for row in CONTROLS if row[3] == "Inherited")])
    summary_ws.append([])
    summary_ws.append(["Remaining gap", "RA-3 (Risk Assessment) -- ISSUE 4.6, GitHub #48, tracked in POAM.md"])
    summary_ws.column_dimensions["A"].width = 24
    summary_ws.column_dimensions["B"].width = 70
    for row_idx in (4, 5):
        summary_ws.cell(row=row_idx, column=1).fill = STATUS_FILLS["Implemented" if row_idx == 4 else "Planned"]
        summary_ws.cell(row=row_idx, column=2).fill = STATUS_FILLS["Implemented" if row_idx == 4 else "Planned"]

    return wb


def main() -> None:
    wb = build_workbook()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Wrote {len(CONTROLS)} controls to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
