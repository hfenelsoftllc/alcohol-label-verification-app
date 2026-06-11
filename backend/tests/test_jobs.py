"""Tests for the batch submission and /jobs/* endpoints."""

import csv
import io
import json

import openpyxl

from app.models import LABEL_FIELD_NAMES, ApplicationData, GovernmentWarningCheck
from app.stubs import build_stub_result
from tests.conftest import PNG_1X1

#: A row of application data that satisfies every required ApplicationData field.
VALID_APPLICATION_ROW = {
    "brand": "Stone's Throw",
    "class_type": "Vodka",
    "abv": "40% Alc. by Vol.",
    "net_contents": "750 mL",
    "name_address": "Stone's Throw Distillery, Louisville, KY",
    "country_of_origin": "United States",
    "government_warning": (
        "GOVERNMENT WARNING: (1) According to the Surgeon General, women "
        "should not drink alcoholic beverages during pregnancy because of "
        "the risk of birth defects. (2) Consumption of alcoholic beverages "
        "impairs your ability to drive a car or operate machinery, and may "
        "cause health problems."
    ),
}


def application_csv(n_rows: int, row: dict | None = None) -> bytes:
    """Build a valid application_csv with `n_rows` identical rows."""
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=LABEL_FIELD_NAMES)
    writer.writeheader()
    for _ in range(n_rows):
        writer.writerow(row or VALID_APPLICATION_ROW)
    return buffer.getvalue().encode("utf-8")


def _submit_batch(client, n_images=3):
    files = [("images", (f"label_{i}.png", PNG_1X1, "image/png")) for i in range(n_images)]
    files.append(("application_csv", ("data.csv", application_csv(n_images), "text/csv")))
    return client.post("/verify/batch", files=files)


def test_batch_returns_job_id(client):
    resp = _submit_batch(client, n_images=3)
    assert resp.status_code == 202
    body = resp.json()
    assert body["total"] == 3
    assert body["state"] == "PENDING"
    assert body["job_id"]


def test_batch_rejects_non_image_415(client):
    files = [("images", ("notes.txt", b"hello", "text/plain"))]
    files.append(("application_csv", ("data.csv", b"brand\nX", "text/csv")))
    resp = client.post("/verify/batch", files=files)
    assert resp.status_code == 415


def test_job_status_results_export_roundtrip(client):
    job_id = _submit_batch(client).json()["job_id"]

    status_resp = client.get(f"/jobs/{job_id}/status")
    assert status_resp.status_code == 200
    assert status_resp.json()["total"] == 3

    results_resp = client.get(f"/jobs/{job_id}/results")
    assert results_resp.status_code == 200
    assert results_resp.json()["summary"]["match"] == 0  # nothing processed yet

    export_resp = client.get(f"/jobs/{job_id}/export")
    assert export_resp.status_code == 200
    assert export_resp.headers["content-type"].startswith("text/csv")
    header = export_resp.text.splitlines()[0].split(",")
    assert header[:4] == ["filename", "overall_status", "confidence_score", "image_quality_score"]
    for name in ("brand_extracted", "brand_expected", "brand_status", "brand_confidence"):
        assert name in header
    assert "government_warning_extracted" in header
    assert "government_warning_expected" in header
    assert "government_warning_status" in header


def test_job_export_xlsx_format(client):
    job_id = _submit_batch(client, n_images=1).json()["job_id"]
    client.get(f"/jobs/{job_id}/stream")  # drive the batch to completion

    resp = client.get(f"/jobs/{job_id}/export?format=xlsx")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert resp.headers["content-disposition"] == f'attachment; filename="results_{job_id}.xlsx"'

    workbook = openpyxl.load_workbook(io.BytesIO(resp.content))
    sheet = workbook.active
    header = [cell.value for cell in sheet[1]]
    assert header[:4] == ["filename", "overall_status", "confidence_score", "image_quality_score"]
    assert "government_warning_status" in header

    # Header row uses the Treasury brand color with bold white text.
    header_cell = sheet.cell(row=1, column=1)
    assert header_cell.fill.fgColor.rgb == "FF0B5D44"
    assert header_cell.font.bold is True

    # The overall_status cell is colour-coded for its status value.
    overall_col = header.index("overall_status") + 1
    status_cell = sheet.cell(row=2, column=overall_col)
    assert status_cell.fill.fgColor.rgb in ("FFC6EFCE", "FFFFEB9C", "FFFFC7CE")


def test_job_export_invalid_format_422(client):
    job_id = _submit_batch(client, n_images=1).json()["job_id"]
    resp = client.get(f"/jobs/{job_id}/export?format=pdf")
    assert resp.status_code == 422


def test_label_result_memory_footprint(application_data):
    """ISSUE 3.5 AC: ~2KB per result / ~600KB for a 300-label batch — a
    realistic VerificationResult (incl. Government Warning text) serializes
    to a small, bounded size, well within in-memory limits even for the
    largest supported batch."""
    app_data = ApplicationData(**application_data)
    result = build_stub_result(app_data, filename="label_001.png")
    result.government_warning = GovernmentWarningCheck(
        valid=True,
        extracted_text=app_data.government_warning,
        expected_text=app_data.government_warning,
    )

    size = len(json.dumps(result.model_dump(mode="json")))

    assert size < 4096
    assert size * 300 < 2 * 1024 * 1024


def test_unknown_job_404(client):
    assert client.get("/jobs/does-not-exist/status").status_code == 404
    assert client.get("/jobs/does-not-exist/results").status_code == 404
    assert client.get("/jobs/does-not-exist/export").status_code == 404
    assert client.get("/jobs/does-not-exist/stream").status_code == 404


def test_openapi_docs_available(client):
    assert client.get("/openapi.json").status_code == 200


def test_batch_populates_job_labels(client):
    from batch import store

    job_id = _submit_batch(client, n_images=2).json()["job_id"]
    job = store.get_job(job_id)
    assert len(job.labels) == 2
    assert [label.filename for label in job.labels] == ["label_0.png", "label_1.png"]
    assert job.labels[0].application_data.brand == VALID_APPLICATION_ROW["brand"]


def test_batch_rejects_csv_missing_columns_422(client):
    files = [("images", ("label_0.png", PNG_1X1, "image/png"))]
    files.append(("application_csv", ("data.csv", b"brand,abv\nX,40", "text/csv")))
    resp = client.post("/verify/batch", files=files)
    assert resp.status_code == 422


def test_batch_rejects_csv_row_count_mismatch_422(client):
    files = [("images", (f"label_{i}.png", PNG_1X1, "image/png")) for i in range(2)]
    files.append(("application_csv", ("data.csv", application_csv(1), "text/csv")))
    resp = client.post("/verify/batch", files=files)
    assert resp.status_code == 422


def test_batch_rejects_csv_invalid_value_422(client):
    bad_row = dict(VALID_APPLICATION_ROW, brand="x" * 300)  # exceeds max_length=255
    files = [("images", ("label_0.png", PNG_1X1, "image/png"))]
    files.append(("application_csv", ("data.csv", application_csv(1, row=bad_row), "text/csv")))
    resp = client.post("/verify/batch", files=files)
    assert resp.status_code == 422
